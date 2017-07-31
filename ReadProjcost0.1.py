from openpyxl import load_workbook
from openpyxl import Workbook
import locale
from tkinter import *

#Richard Jolly 2017
#Version 0.1

root = Tk()
root.title("Test Window")

#Set Locale
locale.setlocale( locale.LC_ALL, '' )

#Open Project XL File holding budget information
xlfile = (r"M:\IT\Change\Current Projects\Windows Server 2003 (F185848)\0 - Project Controls\Costing Budgets\Cost Tracker v0.1.xlsx")

#Load the workbook into Python openpyxl
wb = load_workbook(filename = xlfile, data_only=True)

#Set Active Worksheet as "Costs" Worksheet
ws = wb.active
ws.title = "Costs"

#Retireve calculated vlaue in cell F63
total=ws['F63'].value

#Convert raw date to UK Currency
total=locale.currency(total, grouping=True)

#Save and Close Workbook
wb.save(filename = xlfile)

#Button(root, text="Press here", command=total).pack()

#Output Budget amount to screen
print(total)

root.mainloop()

#Version changes
